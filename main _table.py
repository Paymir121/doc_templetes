import glob
import re

from docxtpl import DocxTemplate
from openpyxl import load_workbook
from translate import translate


def read_excel():
    wb = load_workbook('./export.xlsx')
    ws = wb['Sheet 1']
    table_contents = []

    for id, row in enumerate(ws.iter_rows(min_row=2, max_col=7)):
        name = row[0].value
        phase = row[1].value
        responsible = row[2].value
        document = row[5].value
        step = row[6].value
        table_contents.append({
            'NameRUS': name,
            'PhaseRUS': phase,
            'ResponsibleRUS': responsible,
            'DocumentRUS': document,
            'StepRUS': step,
            'NameENG': translate(name),
            'PhaseENG': translate(phase),
            'ResponsibleENG': translate(responsible),
            'DocumentENG': translate(document),
            'StepENG': translate(step),
        })
    return table_contents


def render_word_file(context):
    docs_files = glob.glob('template.docx') # Читаем все имена файлов *.docx
    for file_name in docs_files:
        if re.search(r'filled_', file_name) is None:
            print(f"<----------------------Делаем красиво с  {file_name}----------------------------------->")
            doc = DocxTemplate(file_name) # Читаем файл
            print(context)
            doc.render(context) # Производим замену name на data
            doc.save('filled_'+file_name) # Сохраняем файл с именем filled_{file_name}
            print(f"<----------------------Сделано красиво с  {file_name}---------------------------------->")


if __name__ == '__main__':
    print("<----------------------Погнали----------------------------------->")
    table_contents = read_excel()
    context = {
        "table_contents": table_contents,
    }
    render_word_file(context)
    print("<----------------------Конец-------------------------------------->")